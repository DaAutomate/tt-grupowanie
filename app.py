import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
import random

st.set_page_config(page_title="Grupowanie Fraz", layout="wide")
st.title("Analiza i Grupowanie Fraz")

# File uploader
uploaded_file = st.file_uploader("Wybierz plik CSV", type=['csv'])

if uploaded_file is not None:
    # Read CSV file
    df = pd.read_csv(uploaded_file)
    
    # Create groups based on URL similarities
    def create_groups(df):
        groups = []
        used_indices = set()
        
        for i in range(len(df)):
            if i in used_indices:
                continue
                
            current_group = [i]
            # Zbierz wszystkie URL-e z wiersza i
            urls_i = set([
                df.iloc[i]['url google 1'],
                df.iloc[i]['url google 2'],
                df.iloc[i]['url google 3']
            ])
            
            for j in range(i + 1, len(df)):
                if j in used_indices:
                    continue
                    
                # Zbierz wszystkie URL-e z wiersza j
                urls_j = set([
                    df.iloc[j]['url google 1'],
                    df.iloc[j]['url google 2'],
                    df.iloc[j]['url google 3']
                ])
                
                # Sprawdź czy jest jakiś wspólny URL
                if len(urls_i.intersection(urls_j)) > 0:
                    current_group.append(j)
                    used_indices.add(j)
            
            if len(current_group) > 1:  # Dodaj grupę tylko jeśli ma więcej niż jeden element
                groups.append(current_group)
            used_indices.add(i)
            
        return groups
    
    # Process data
    groups = create_groups(df)
    
    # Assign group numbers and main topics
    df['numer grupy'] = -1
    df['główny temat'] = ''
    
    # Najpierw przypisz numery grup dla powiązanych fraz
    for idx, group in enumerate(groups, 1):
        for i in group:
            df.loc[i, 'numer grupy'] = idx
    
    # Następnie ustal główne tematy dla każdej grupy
    for group_number in df[df['numer grupy'] != -1]['numer grupy'].unique():
        # Wybierz wszystkie frazy z danej grupy
        group_df = df[df['numer grupy'] == group_number]
        
        # Znajdź frazę z największym volumenem
        max_vol_idx = group_df['Vol'].idxmax()
        main_topic = df.loc[max_vol_idx, 'KW']
        
        # Przypisz główny temat wszystkim frazom w grupie
        df.loc[df['numer grupy'] == group_number, 'główny temat'] = main_topic

    # Obsłuż frazy bez grupy
    next_group_number = len(groups) + 1 if groups else 1
    mask_no_group = df['numer grupy'] == -1
    
    # Dla każdej frazy bez grupy:
    # - ustaw jej własną frazę jako główny temat
    # - przypisz kolejny numer grupy
    df.loc[mask_no_group, 'główny temat'] = df.loc[mask_no_group, 'KW']
    df.loc[mask_no_group, 'numer grupy'] = range(next_group_number, next_group_number + mask_no_group.sum())

    # Display results with colors
    st.subheader("Wyniki grupowania")
    
    # Create a color map for groups
    unique_groups = df[df['numer grupy'] != -1]['numer grupy'].unique()
    color_map = {group: f"#{random.randint(0, 0xFFFFFF):06x}" for group in unique_groups}
    
    # Style the dataframe
    def highlight_rows(row):
        if row['numer grupy'] != -1:
            color = color_map[row['numer grupy']]
            return [f'background-color: {color}'] * len(row)
        return [''] * len(row)
    
    styled_df = df.style.apply(highlight_rows, axis=1)
    st.dataframe(styled_df)

    # Export button
    if st.button("Zapisz wyniki do Excel"):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Wyniki')
            workbook = writer.book
            worksheet = writer.sheets['Wyniki']
            
            # Color coding for groups
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1)):
                group_num = df.iloc[idx]['numer grupy']
                if group_num != -1:
                    color = color_map[group_num]
                    fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")
                    for cell in row:
                        cell.fill = fill

        output.seek(0)
        st.download_button(
            label="Pobierz plik Excel",
            data=output.getvalue(),
            file_name="wyniki_grupowania.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Display statistics
    st.subheader("Statystyki")
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("Liczba znalezionych grup", len(groups))
    
    with col2:
        phrases_without_group = len(df[df['numer grupy'] == -1])
        st.metric("Liczba fraz bez grupy", phrases_without_group)
